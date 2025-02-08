from flask import Flask, render_template, request
import mysql.connector
import datetime
import os
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate, migrate

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://' + os.getenv('MYSQL_USER') + ':' + os.getenv('MYSQL_PASSWORD') + '@' + os.getenv('MYSQL_HOST') + '/' + os.getenv('MYSQL_DATABASE')

db = SQLAlchemy(app)
migrate = Migrate(app, db)

def get_db_connection():
    connection = mysql.connector.connect(
        host=os.getenv('MYSQL_HOST'),
        user=os.getenv('MYSQL_USER'),
        password=os.getenv('MYSQL_PASSWORD'),
        database=os.getenv('MYSQL_DATABASE')
    )
    return connection

@app.route('/', methods=['GET', 'POST'])
def index():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute('SELECT DATABASE()')
        sqldb = cursor.fetchone()
        try:
            if request.method == 'POST':
                info = []
                # if 'file' not in request.files:
                #     return 'No file part'
                uploaded_files = request.files.getlist('file[]')
                if uploaded_files[1] is not None:
                    for file in uploaded_files:
                    # file = request.files['file']

                        if file.filename == '':
                            return 'No selected file'
                        if file.filename.endswith('.pdf'):
                            return 'Not excel'
                        wb = load_workbook(file)
                        first_num = None
                        last_num = None
                        date_string = None
                        for index, char in enumerate(file.filename):
                            if char.isdigit():
                                if first_num is None:
                                    first_num = index
                                last_num = index
                        date_list = file.filename[first_num:last_num + 1].split(" ")
                        date_obj = datetime.datetime(int("20" + date_list[2]), int(date_list[0]), int(date_list[1]))
                        for index, item in enumerate(date_list):
                            if len(item) < 2:
                                date_list[index] = item.zfill(2)
                        
                        date_string = "-".join(date_list)

                        # info.append(date_obj.strftime("%m %d %Y"))
                        # dates need fixing if not in mm dd yy, many in m d yy
                        # class paymentRow():
                        #     def __init__(self, person_id, name, rent_paid, food_paid, rent_balance, food_balance, rent_carry, food_carry, date):
                        #         self.person_id = person_id
                        #         self.name = name
                        #         self.rent_paid = rent_paid
                        #         self.food_paid = food_paid
                        #         self.rent_balance = rent_balance 
                        #         self.food_balance = food_balance
                        #         self.rent_carry = rent_carry
                        #         self.food_carry = food_carry
                        #         self.date = date
                        for row in wb.active.iter_rows(3, 13, 2, 12, True):
                            person = Person.query.filter_by(name=row[0]).first()
                            if person is None and row[0] is not None:
                                        
                                person = Person(name=row[0])
                                db.session.add(person)
                                db.session.commit()
                            person = Person.query.filter_by(name=row[0]).first()
                            # row[4] rent paid row[7] equation/Balance row[9] carry
                            thisPaymentRow = PaymentRow.query.filter_by(date = date_obj.strftime("%Y-%m-%d"), name = row[0]).first()
                            # if(row[0] is not None and thisPaymentRow is None):
                            if thisPaymentRow is None and row[0] is not None:
                                row = list(row)
                                house_position = row[6] or "None"
                                rent_cost = int(row[7][row[7].find("-")+1:])
                                food_cost = int(row[8][row[8].find("-")+1:])
                                rent_paid = row[4] or 0
                                rent_carry = row[9] or 0
                                rent_balance = rent_paid + rent_carry - rent_cost
                                food_paid = row[5] or 0
                                food_carry = row[10] or 0
                                food_balance = food_paid + food_carry - food_cost
                                # thisPaymentRow = PaymentRow(person.id, row[0], rent_paid, food_paid, rent_balance, food_balance, rent_carry, food_carry, date_obj.strftime("%m %d %Y"))
                                thisPaymentRow = PaymentRow(person_id = person.id, name = row[0], rent_paid = rent_paid, food_paid = food_paid, rent_balance = rent_balance, food_balance = food_balance, rent_carry = rent_carry, food_carry = food_carry, date = date_obj.strftime("%Y-%m-%d"))                            
                                db.session.add(thisPaymentRow)
                                db.session.commit()
                            info.append(thisPaymentRow)
                                
                                # for cell in row:
                                    # info.append(cell)
                                    # info.append(date_string)
                                    # person = Person.query.filter_by(name=cell).first()
                                    # if person is None and cell is not None:
                                        
                                        
                                        # person = Person(name=cell)
                                        # db.session.add(person)
                                        # db.session.commit()
                        # info = wb.active['B3'].value
                            
                return render_template('index.html', info = info)
            info = PaymentRow.query.all()
            return render_template('index.html', info=info, database=sqldb[0])
        except Exception as e:
            return render_template('index.html', error=str(e))
    except Exception as e:
        return render_template('index.html', error=str(e))
    finally:
        if 'connection' in locals():
            connection.close()
    

class Person(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=False, nullable=False)
    notes = db.Column(db.Text(1024), unique=False, nullable=True)

    def __repr__(self):
        return f"Name : {self.name}"

class PaymentRow(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    person_id = db.Column(db.Integer, db.ForeignKey('person.id'), unique=False, nullable=False)
    name = db.Column(db.String(120), unique=False, nullable=False)
    house_position = db.Column(db.String(120), unique=False, nullable=True)
    rent_paid = db.Column(db.Float, unique=False, nullable=False)
    food_paid = db.Column(db.Float, unique=False, nullable=False)
    rent_balance = db.Column(db.Float, unique=False, nullable=False)
    food_balance = db.Column(db.Float, unique=False, nullable=False)
    rent_carry = db.Column(db.Float, unique=False, nullable=False)
    food_carry = db.Column(db.Float, unique=False, nullable=False)
    date = db.Column(db.Date, unique=False, nullable=False)
    notes = db.Column(db.Text(1024), unique=False, nullable=True)
    def __repr__(self):
        return f"Name : {self.name} Date: {self.date}"


# Settings for migrations



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)

@app.route('/init')
def init():
    db.create_all()
    return '<h1>Inititted</h1>'